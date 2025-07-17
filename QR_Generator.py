import os
import string
import itertools
from datetime import datetime
import qrcode
from PIL import Image, ImageDraw, ImageFont
import openpyxl

# === Configuration ===
OUTPUT_DIR = "qr_output"
EXCEL_FILE = r"C:\Git\QR Codes\qr_codes.xlsx"
FONT_PATH = "C:/Windows/Fonts/arial.ttf"  # Adjust if needed
FONT_SIZE = 80
QR_SIZE = 300

# === Ensure output folder exists ===
os.makedirs(OUTPUT_DIR, exist_ok=True)

# === Load existing codes from Excel ===
def load_existing_serials():
    serials = set()
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
        sheet = wb.get_sheet_by_name("Generated") if "Generated" in wb.sheetnames else None
        if sheet:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                serials.add(row[0])
    print(f"[INFO] Found {len(serials)} existing codes in 'Generated' sheet.")
    return serials

# === Save new codes to Excel ===
def save_to_excel(new_data):
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet if it exists
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # Create or access "Generated" sheet
    if "Generated" not in wb.sheetnames:
        sheet = wb.create_sheet("Generated")
        sheet.append(["Serial", "Generated At", "Note"])
    else:
        sheet = wb["Generated"]

    for data in new_data:
        sheet.append(data)

    wb.save(EXCEL_FILE)
    print(f"[INFO] {len(new_data)} new QR codes appended to {EXCEL_FILE} ('Generated' sheet).")

# === Generator for serials: fixed range ===
def serial_generator(start: str, count: int = None):
    charset = string.ascii_uppercase
    length = len(start)
    all_serials = (''.join(p) for p in itertools.product(charset, repeat=length))

    skipping = True
    generated = 0
    for serial in all_serials:
        if skipping:
            if serial == start:
                skipping = False
            else:
                continue

        yield serial
        generated += 1
        if count is not None and generated >= count:
            break

# === Open-ended serial generator (Mode 2) ===
def serial_generator_open_ended(start: str):
    charset = string.ascii_uppercase
    length = len(start)
    all_serials = (''.join(p) for p in itertools.product(charset, repeat=length))

    skipping = True
    for serial in all_serials:
        if skipping:
            if serial == start:
                skipping = False
            else:
                continue
        yield serial

# === Generate a QR code with text below ===
def generate_qr_with_text(serial: str, size: int = QR_SIZE):
    qr = qrcode.QRCode(border=1)
    qr.add_data(serial)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
    img = img.resize((size, size), Image.Resampling.LANCZOS)

    # Load font safely
    try:
        font = ImageFont.truetype(FONT_PATH, FONT_SIZE)
    except OSError:
        print(f"[WARNING] Font not found at {FONT_PATH}, using default font.")
        font = ImageFont.load_default()

    # Measure text size
    text = serial
    bbox = font.getbbox(text)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]

    # Make canvas wide enough for both QR and text
    canvas_width = max(size, text_width + 20)  # Add padding
    canvas_height = size + text_height + 10

    canvas = Image.new("RGB", (canvas_width, canvas_height), "white")

    # Paste QR in the center
    qr_x = (canvas_width - size) // 2
    canvas.paste(img, (qr_x, 0))

    # Draw text centered below
    draw = ImageDraw.Draw(canvas)
    text_x = (canvas_width - text_width) // 2
    draw.text((text_x, size-10), text, font=font, fill="black")

    filename = os.path.join(OUTPUT_DIR, f"{serial}.png")
    canvas.save(filename)
    print(f"[INFO] Saved QR code: {filename}")

# === Main logic ===
def main():
    print("=== QR Code Batch Generator ===")
    print("Choose mode (1=range AAA to AAZ, 2=start BAA with count 100, 3=read from 'Create' sheet): ", end="")
    mode = input().strip()

    existing_serials = load_existing_serials()
    new_data = []

    if mode == "1":
        print("Enter starting serial (e.g. AAA): ", end="")
        start = input().strip().upper()
        print("Enter ending serial (e.g. AAZ): ", end="")
        end = input().strip().upper()
        print("Enter note to save with each code (optional): ", end="")
        note = input().strip()

        generator = serial_generator(start)
        for serial in generator:
            if serial > end:
                break
            if serial in existing_serials:
                print(f"[SKIP] Serial {serial} already exists.")
                continue
            generate_qr_with_text(serial)
            new_data.append((serial, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), note))

    elif mode == "2":
        print("Enter starting serial (e.g. BAA): ", end="")
        start = input().strip().upper()
        print("Enter how many serials to generate: ", end="")
        try:
            count = int(input().strip())
        except ValueError:
            print("[ERROR] Invalid count.")
            return
        print("Enter note to save with each code (optional): ", end="")
        note = input().strip()

        generator = serial_generator_open_ended(start)
        max_attempts = 10000
        attempts = 0
        while len(new_data) < count:
            serial = next(generator)
            attempts += 1

            if serial in existing_serials:
                print(f"[SKIP] Serial {serial} already exists.")
                continue

            generate_qr_with_text(serial)
            new_data.append((serial, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), note))

            if attempts >= max_attempts:
                print("[ERROR] Too many attempts. Could not generate enough unique serials.")
                break

    elif mode == "3":
        if not os.path.exists(EXCEL_FILE):
            print(f"[ERROR] Excel file {EXCEL_FILE} does not exist.")
            return

        try:
            wb = openpyxl.load_workbook(EXCEL_FILE)
        except Exception as e:
            print(f"[ERROR] Failed to open {EXCEL_FILE}: {e}")
            return

        if "Create" not in wb.sheetnames:
            print("[ERROR] Sheet 'Create' not found in the Excel file.")
            return

        sheet = wb["Create"]
        headers = [cell.value for cell in sheet[1]]
        if headers[:2] != ["Serial", "Note"]:
            print("[ERROR] 'Create' sheet must have headers 'Serial' and 'Note' in columns A and B.")
            return

        for row in sheet.iter_rows(min_row=2, values_only=True):
            serial = str(row[0]).strip().upper() if row[0] else None
            note = str(row[1]).strip() if row[1] else ""

            if not serial:
                print("[SKIP] Empty serial found in 'Create' sheet.")
                continue

            if serial in existing_serials:
                print(f"[SKIP] Serial {serial} already exists in 'Generated' sheet.")
                continue

            generate_qr_with_text(serial)
            new_data.append((serial, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), note))

    else:
        print("[ERROR] Invalid mode selected.")
        return

    if new_data:
        try:
            save_to_excel(new_data)
        except Exception as e:
            print(f"[ERROR] Failed to save to {EXCEL_FILE}: {e}")
    else:
        print("[INFO] No new codes were generated.")

if __name__ == "__main__":
    main()
